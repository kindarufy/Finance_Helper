"""Модуль сервисного слоя Telegram-бота Finance Helper."""
from __future__ import annotations

import csv
import hashlib
import re
from datetime import date, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


def parse_add_command(text: str):
    """Разбирает входные данные для сценария «add command»."""
    parts = text.split()
    if len(parts) < 4:
        return None
    amount_raw = parts[1]
    type_raw = parts[2].lower()
    category = parts[3]
    comment = " ".join(parts[4:]) if len(parts) > 4 else None
    try:
        amount = float(amount_raw.replace(",", "."))
        if amount <= 0:
            return None
    except Exception:
        return None
    if type_raw == "расход":
        op_type = "expense"
    elif type_raw == "доход":
        op_type = "income"
    else:
        return None
    return amount, op_type, category, comment


def parse_report(text: str):
    """Разбирает входные данные для сценария «report»."""
    parts = text.split()
    if len(parts) != 3:
        return None
    d1 = parts[1].strip()
    d2 = parts[2].strip()
    try:
        date.fromisoformat(d1)
        date.fromisoformat(d2)
    except Exception:
        return None
    return d1, d2


_DATE_EXACT_PATTERNS = [
    re.compile(r"^сегодня$", flags=re.IGNORECASE),
    re.compile(r"^вчера$", flags=re.IGNORECASE),
    re.compile(r"^позавчера$", flags=re.IGNORECASE),
    re.compile(r"^(\d{1,3})\s+д(?:ень|ня|ней)\s+назад$", flags=re.IGNORECASE),
    re.compile(r"^(\d{4}-\d{2}-\d{2})$"),
    re.compile(r"^(\d{1,2})\.(\d{1,2})(?:\.(\d{2,4}))?$"),
]


def parse_user_date(value: str, today: date | None = None) -> date | None:
    """Разбирает входные данные для сценария «user date»."""
    today = today or date.today()
    cleaned = " ".join((value or "").strip().split())
    if not cleaned:
        return None
    if _DATE_EXACT_PATTERNS[0].match(cleaned):
        return today
    if _DATE_EXACT_PATTERNS[1].match(cleaned):
        return today - timedelta(days=1)
    if _DATE_EXACT_PATTERNS[2].match(cleaned):
        return today - timedelta(days=2)
    back_match = _DATE_EXACT_PATTERNS[3].match(cleaned)
    if back_match:
        return today - timedelta(days=int(back_match.group(1)))
    iso_match = _DATE_EXACT_PATTERNS[4].match(cleaned)
    if iso_match:
        try:
            return date.fromisoformat(iso_match.group(1))
        except ValueError:
            return None
    ru_match = _DATE_EXACT_PATTERNS[5].match(cleaned)
    if ru_match:
        try:
            day = int(ru_match.group(1))
            month = int(ru_match.group(2))
            year_raw = ru_match.group(3)
            if year_raw is None:
                year = today.year
            else:
                year = int(year_raw)
                if year < 100:
                    year += 2000
            return date(year, month, day)
        except ValueError:
            return None
    return None


def _extract_date_fragment(text: str, today: date) -> tuple[date | None, str]:
    """Выполняет действие «extract date fragment» в рамках логики Finance Helper."""
    cleaned = text.strip()
    for token, dt in {"сегодня": today, "вчера": today - timedelta(days=1), "позавчера": today - timedelta(days=2)}.items():
        if re.search(rf"(?:^|\s){token}(?:$|\s)", cleaned, flags=re.IGNORECASE):
            cleaned = re.sub(rf"(?:^|\s){token}(?:$|\s)", " ", cleaned, flags=re.IGNORECASE).strip()
            return dt, " ".join(cleaned.split())
    back_match = re.search(r"(?:^|\s)(\d{1,3})\s+д(?:ень|ня|ней)\s+назад(?:$|\s)", cleaned, flags=re.IGNORECASE)
    if back_match:
        cleaned = cleaned.replace(back_match.group(0), " ").strip()
        return today - timedelta(days=int(back_match.group(1))), " ".join(cleaned.split())
    iso_match = re.search(r"(\d{4}-\d{2}-\d{2})", cleaned)
    if iso_match:
        try:
            dt = date.fromisoformat(iso_match.group(1))
            cleaned = cleaned.replace(iso_match.group(1), " ").strip()
            return dt, " ".join(cleaned.split())
        except ValueError:
            pass
    ru_match = re.search(r"(\d{1,2})\.(\d{1,2})(?:\.(\d{2,4}))?", cleaned)
    if ru_match:
        try:
            day_value = int(ru_match.group(1))
            month_value = int(ru_match.group(2))
            year_raw = ru_match.group(3)
            if year_raw is None:
                year = today.year
            else:
                year = int(year_raw)
                if year < 100:
                    year += 2000
            dt = date(year, month_value, day_value)
            cleaned = cleaned.replace(ru_match.group(0), " ").strip()
            return dt, " ".join(cleaned.split())
        except ValueError:
            return None, " ".join(cleaned.split())
    return None, " ".join(cleaned.split())


_AMOUNT_RE = re.compile(r"^[+]?\d+(?:[.,]\d+)?$")
_CURRENCY_RE = re.compile(r"^[A-Za-z]{3,5}$")


def parse_natural_operation(text: str, today: date | None = None) -> dict[str, Any] | None:
    """Разбирает входные данные для сценария «natural operation»."""
    if not text:
        return None
    source = " ".join(text.strip().split())
    if not source or source.startswith("/"):
        return None
    tokens = source.split()
    first = tokens[0]
    if not _AMOUNT_RE.match(first):
        return None
    amount = float(first.replace(",", ".").lstrip("+"))
    if amount <= 0:
        return None
    op_type = "income" if first.startswith("+") else "expense"
    rest_tokens = tokens[1:]
    currency = "RUB"
    if rest_tokens and _CURRENCY_RE.match(rest_tokens[0]):
        currency = rest_tokens[0].upper()
        rest_tokens = rest_tokens[1:]
    rest_text = " ".join(rest_tokens)
    occurred_at, rest_text = _extract_date_fragment(rest_text, today or date.today())
    comment = rest_text.strip() or None
    return {
        "amount": amount,
        "op_type": op_type,
        "currency": currency,
        "occurred_at": occurred_at,
        "raw_text": source,
        "description": comment,
    }


DEFAULT_CATEGORY_HINTS: dict[str, dict[str, set[str]]] = {
    "expense": {
        "Еда": {"еда", "пицца", "кофе", "обед", "ужин", "кафе", "ресторан", "продукты", "суши", "бургер", "магнит", "пятерочка"},
        "Транспорт": {"такси", "метро", "автобус", "транспорт", "бензин", "uber", "bolt", "поезд"},
        "Дом": {"аренда", "квартира", "дом", "коммунал", "ремонт", "мебель", "икеа"},
        "Развлечения": {"кино", "бар", "концерт", "театр", "игра", "музей", "вечеринка"},
        "Здоровье": {"аптека", "лекар", "врач", "здоров", "стоматолог"},
        "Образование": {"курс", "книга", "школа", "урок", "обучение", "универ", "образование"},
    },
    "income": {
        "Зарплата": {"зарплата", "зп", "salary", "оклад"},
        "Премия": {"премия", "bonus", "бонус"},
        "Продажа вещей": {"авито", "продажа", "продал", "продала", "вещи"},
        "Подарки": {"подарок", "подарили", "дарение"},
    },
}


def infer_default_category(description: str | None, op_type: str) -> str | None:
    """Выполняет действие «infer default category» в рамках логики Finance Helper."""
    if not description:
        return None
    normalized = f" {description.lower()} "
    for category, hints in DEFAULT_CATEGORY_HINTS.get(op_type, {}).items():
        for hint in hints:
            if f" {hint} " in normalized or hint in normalized:
                return category
    return None


_CURRENCY_SYMBOLS = {"₽": "RUB", "$": "USD", "€": "EUR", "₸": "KZT", "£": "GBP"}


def _detect_currency(text: str) -> str:
    """Выполняет действие «detect currency» в рамках логики Finance Helper."""
    upper = text.upper()
    for code in ["RUB", "USD", "EUR", "KZT", "GBP"]:
        if code in upper:
            return code
    if re.search(r"\b(?:РУБ|РУБЛ|РУБЛЕЙ|RUR)\b", upper):
        return "RUB"
    if re.search(r"\b(?:ДОЛЛАР|USD)\b", upper):
        return "USD"
    if re.search(r"\b(?:ЕВРО|EUR)\b", upper):
        return "EUR"
    for symbol, code in _CURRENCY_SYMBOLS.items():
        if symbol in text:
            return code
    return "RUB"


def _parse_any_date_fragment(value: str, today: date | None = None) -> date | None:
    """Выполняет действие «parse any date fragment» в рамках логики Finance Helper."""
    value = (value or "").strip()
    if not value:
        return None
    parsed = parse_user_date(value, today=today or date.today())
    if parsed is not None:
        return parsed
    m = re.search(r"(\d{1,2})[\./-](\d{1,2})[\./-](\d{2,4})", value)
    if m:
        day_value = int(m.group(1)); month_value = int(m.group(2)); year = int(m.group(3))
        if year < 100:
            year += 2000
        try:
            return date(year, month_value, day_value)
        except ValueError:
            return None
    return None


def _receipt_candidate_lines(text: str) -> list[str]:
    """Выполняет действие «receipt candidate lines» в рамках логики Finance Helper."""
    return [line.strip() for line in text.splitlines() if line and line.strip()]


def _extract_receipt_items(text: str) -> list[dict[str, Any]]:
    """Выполняет действие «extract receipt items» в рамках логики Finance Helper."""
    items: list[dict[str, Any]] = []
    for line in _receipt_candidate_lines(text):
        if len(items) >= 5:
            break
        m = re.search(r"([A-Za-zА-Яа-я0-9 .,_\-]{3,})\s+(\d+[\s\d]*[\.,]\d{2}|\d{2,})$", line)
        if not m:
            continue
        label = " ".join(m.group(1).split())
        if re.search(r"итог|сумма|к оплате|cash|change|скидка", label, flags=re.IGNORECASE):
            continue
        amount_raw = m.group(2).replace(" ", "").replace(",", ".")
        try:
            amount = float(amount_raw)
        except Exception:
            continue
        if amount <= 0:
            continue
        items.append({"label": label[:80], "amount": round(amount, 2)})
    return items


def _best_merchant_from_lines(lines: list[str], fallback: str | None = None) -> str | None:
    """Выполняет действие «best merchant from lines» в рамках логики Finance Helper."""
    for line in lines[:8]:
        if len(line) < 3:
            continue
        if re.search(r"\d", line):
            continue
        if re.search(r"итог|сумма|касса|чек|спасибо", line, flags=re.IGNORECASE):
            continue
        return line[:255]
    return fallback[:255] if fallback else None


def _preprocessed_ocr_text(image_path: str | Path) -> tuple[str, str | None]:
    """Выполняет действие «preprocessed ocr text» в рамках логики Finance Helper."""
    try:
        from PIL import Image, ImageFilter, ImageOps  # type: ignore
        import pytesseract  # type: ignore
        image = Image.open(str(image_path)).convert("L")
        prepared = [image, ImageOps.autocontrast(image), ImageOps.autocontrast(image).filter(ImageFilter.SHARPEN)]
        chunks: list[str] = []
        for variant in prepared:
            try:
                raw = pytesseract.image_to_string(variant, lang="rus+eng", config="--psm 6")
            except Exception:
                raw = ""
            if raw and raw.strip():
                chunks.append(raw)
        return "\n".join(chunks).strip(), None
    except Exception as exc:  # pragma: no cover
        return "", str(exc)


def extract_receipt_data(image_path: str | Path, hint_text: str | None = None) -> dict[str, Any]:
    """Выполняет действие «extract receipt data» в рамках логики Finance Helper."""
    raw_text_parts: list[str] = []
    if hint_text:
        raw_text_parts.append(hint_text)
    ocr_text, ocr_error = _preprocessed_ocr_text(image_path)
    if ocr_text:
        raw_text_parts.append(ocr_text)
    text = "\n".join(part for part in raw_text_parts if part).strip()
    lines = _receipt_candidate_lines(text)

    priority = re.findall(r"(?:ИТОГ|ИТОГО|К\s*ОПЛАТЕ|СУММА|TOTAL)\D{0,18}(\d+[\s\d]*[\.,]\d{2}|\d+[\s\d]*)", text, flags=re.IGNORECASE)
    scrubbed_text = re.sub(r"\d{4}-\d{2}-\d{2}|\d{1,2}[./]\d{1,2}(?:[./]\d{2,4})?", " ", text)
    moneyish = re.findall(r"\d+[\s\d]*[\.,]\d{2}|\d{2,}[\s\d]*", scrubbed_text)
    amount_candidates: list[float] = []
    for token in priority + moneyish:
        cleaned = token.replace(" ", "").replace(",", ".")
        try:
            value = float(cleaned)
        except Exception:
            continue
        if value > 0:
            amount_candidates.append(value)
    amount = max(amount_candidates) if amount_candidates else None

    merchant = _best_merchant_from_lines(lines, fallback=hint_text.strip() if hint_text else None)
    occurred_at = None
    for pattern in [r"\d{4}-\d{2}-\d{2}", r"\d{1,2}\.\d{1,2}(?:\.\d{2,4})?", r"\d{1,2}/\d{1,2}/\d{2,4}"]:
        m = re.search(pattern, text)
        if m:
            occurred_at = _parse_any_date_fragment(m.group(0))
            if occurred_at:
                break
    currency = _detect_currency(text or hint_text or "")
    return {
        "raw_text": text or hint_text or "",
        "amount": amount,
        "currency": currency,
        "merchant": merchant,
        "occurred_at": occurred_at,
        "ocr_error": ocr_error,
        "items": _extract_receipt_items(text),
    }


def _coerce_amount(raw: Any) -> float | None:
    """Выполняет действие «coerce amount» в рамках логики Finance Helper."""
    if raw is None:
        return None
    value = str(raw).strip().replace(" ", "").replace(",", ".")
    if not value:
        return None
    try:
        return float(value)
    except Exception:
        return None


def _statement_row_to_operation(row: dict[str, Any]) -> dict[str, Any] | None:
    """Выполняет действие «statement row to operation» в рамках логики Finance Helper."""
    dt = _parse_any_date_fragment(str(row.get("date") or row.get("дата") or row.get("operation_date") or row.get("transaction_date") or row.get("posted_at") or ""))
    amount = _coerce_amount(row.get("amount") or row.get("sum") or row.get("сумма"))
    debit = _coerce_amount(row.get("debit") or row.get("expense") or row.get("списание"))
    credit = _coerce_amount(row.get("credit") or row.get("income") or row.get("зачисление"))
    if amount is None:
        if debit not in (None, 0):
            amount = -abs(float(debit))
        elif credit not in (None, 0):
            amount = abs(float(credit))
    if dt is None or amount is None or amount == 0:
        return None
    description = str(row.get("description") or row.get("назначение") or row.get("comment") or row.get("details") or row.get("merchant") or row.get("counterparty") or row.get("операция") or "").strip()
    currency = str(row.get("currency") or row.get("валюта") or _detect_currency(description)).upper()
    op_type = str(row.get("type") or row.get("тип") or "").strip().lower()
    if not op_type:
        op_type = "income" if amount > 0 else "expense"
    if op_type in {"расход", "expense", "debit", "списание"}:
        final_type = "expense"
        amount = abs(amount)
    elif op_type in {"доход", "income", "credit", "зачисление"}:
        final_type = "income"
        amount = abs(amount)
    else:
        final_type = "income" if amount > 0 else "expense"
        amount = abs(amount)
    merchant = description[:255] or None
    fingerprint = f"{dt.isoformat()}|{amount:.2f}|{currency}|{description}|{final_type}"
    external_ref = hashlib.sha1(fingerprint.encode("utf-8")).hexdigest()
    return {
        "occurred_at": dt,
        "amount": amount,
        "currency": currency,
        "comment": description or None,
        "merchant": merchant,
        "type": final_type,
        "external_ref": external_ref,
    }


def parse_statement_file(filename: str, content: bytes) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Разбирает входные данные для сценария «statement file»."""
    suffix = Path(filename or "statement").suffix.lower()
    rows: list[dict[str, Any]] = []
    if suffix == ".csv":
        sample = content.decode("utf-8-sig", errors="ignore")
        try:
            dialect = csv.Sniffer().sniff(sample[:2048], delimiters=",;\t")
        except Exception:
            dialect = csv.excel
            dialect.delimiter = ';' if ';' in sample else ','
        reader = csv.reader(StringIO(sample), dialect)
        all_rows = list(reader)
        if not all_rows:
            return [], {"file_type": "csv", "message": "Пустой CSV"}
        headers = [str(x).strip() for x in all_rows[0]]
        normalized = [re.sub(r"\s+", "_", (h or "").strip().lower()) for h in headers]
        for values in all_rows[1:]:
            row = {}
            for i in range(len(values)):
                key = normalized[i] if i < len(normalized) and normalized[i] else (headers[i] if i < len(headers) else f"col_{i}")
                row[key] = values[i]
            parsed = _statement_row_to_operation(row)
            if parsed:
                rows.append(parsed)
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        if load_workbook is None:
            raise RuntimeError("openpyxl не установлен")
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        if not data:
            return [], {"file_type": "xlsx", "message": "Пустой XLSX"}
        headers = [str(x or "").strip() for x in data[0]]
        normalized = [re.sub(r"\s+", "_", (h or "").strip().lower()) for h in headers]
        for values in data[1:]:
            row = {}
            for i in range(len(values)):
                key = normalized[i] if i < len(normalized) and normalized[i] else (headers[i] if i < len(headers) else f"col_{i}")
                row[key] = values[i]
            parsed = _statement_row_to_operation(row)
            if parsed:
                rows.append(parsed)
    else:
        raise ValueError("Поддерживаются только CSV и XLSX")

    expenses = sum(1 for item in rows if item["type"] == "expense")
    incomes = sum(1 for item in rows if item["type"] == "income")
    total_expense = sum(item["amount"] for item in rows if item["type"] == "expense")
    total_income = sum(item["amount"] for item in rows if item["type"] == "income")
    summary = {
        "file_type": suffix.lstrip('.') or 'unknown',
        "rows": len(rows),
        "expenses": expenses,
        "incomes": incomes,
        "total_expense": round(total_expense, 2),
        "total_income": round(total_income, 2),
        "message": f"Найдено операций: {len(rows)}; расходов: {expenses}; доходов: {incomes}",
    }
    return rows, summary
